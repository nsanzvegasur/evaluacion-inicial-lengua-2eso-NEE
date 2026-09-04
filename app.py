from pathlib import Path

import streamlit as st

# =========================================================
# CONFIGURACIÓN Y CAPA VISUAL NNEE
# =========================================================

st.set_page_config(
    page_title="Evaluación Inicial Lengua 2.º ESO — NNEE",
    page_icon="📚",
    layout="centered",
)

st.markdown(
    """
    <style>
    html, body, [class*="css"],
    [data-testid="stAppViewContainer"],
    [data-testid="stAppViewContainer"] * {
        font-family: "Trebuchet MS", Verdana, sans-serif !important;
    }

    [data-testid="stMainBlockContainer"] {
        max-width: 900px;
        padding-top: 2.2rem;
        padding-bottom: 4rem;
    }

    h1 {
        font-size: 2.15rem !important;
        line-height: 1.25 !important;
        margin-bottom: 1rem !important;
    }

    h1::after {
        content: "  ·  NNEE";
        font-size: 0.72em;
        font-weight: 700;
    }

    h2 {
        font-size: 1.75rem !important;
        line-height: 1.35 !important;
        letter-spacing: 0.01em;
        margin-top: 3.4rem !important;
        margin-bottom: 1.4rem !important;
        padding-top: 1.5rem !important;
        border-top: 3px solid #777 !important;
    }

    h3 {
        font-size: 1.45rem !important;
        line-height: 1.4 !important;
        margin-top: 1.4rem !important;
        margin-bottom: 1.2rem !important;
    }

    p, li, label {
        font-size: 1.12rem !important;
        line-height: 1.75 !important;
        letter-spacing: 0.01em;
    }

    [data-testid="stMarkdownContainer"] {
        font-size: 1.12rem !important;
        line-height: 1.8 !important;
    }

    /* La negrita vuelve a ser neutra. El rojo se reserva para lo importante. */
    strong, b {
        color: inherit !important;
        font-weight: 700 !important;
    }

    .pregunta-roja {
        color: #b00020 !important;
        font-weight: 700 !important;
        font-size: 1.12rem !important;
        line-height: 1.7 !important;
        margin-bottom: 0.15rem !important;
    }

    .etiqueta-roja {
        color: #b00020 !important;
        font-weight: 700 !important;
    }

    textarea, input,
    [data-baseweb="select"] {
        font-family: "Trebuchet MS", Verdana, sans-serif !important;
        font-size: 1.12rem !important;
        line-height: 1.6 !important;
    }

    input, textarea {
        min-height: 3.1rem !important;
        padding: 0.8rem 0.9rem !important;
    }

    textarea {
        min-height: 8rem !important;
    }

    [data-testid="stTextInput"],
    [data-testid="stTextArea"],
    [data-testid="stSelectbox"],
    [data-testid="stNumberInput"] {
        margin-bottom: 1.1rem !important;
    }

    button[kind="primary"] {
        font-size: 1.15rem !important;
        min-height: 3.6rem !important;
        margin-top: 2rem !important;
    }

    input:focus, textarea:focus,
    div[data-baseweb="select"]:focus-within {
        outline: 3px solid #555 !important;
        outline-offset: 2px !important;
    }

    /* Resultado final: grande, centrado y con la misma tipografía adaptada. */
    [data-testid="stMetric"] {
        text-align: center !important;
        margin: 1.8rem 0 2.4rem 0 !important;
        padding: 1.2rem 1rem 1.4rem 1rem !important;
    }

    [data-testid="stMetricLabel"] {
        justify-content: center !important;
        font-family: "Trebuchet MS", Verdana, sans-serif !important;
        font-size: 1.25rem !important;
        line-height: 1.4 !important;
        font-weight: 700 !important;
        color: #444 !important;
    }

    [data-testid="stMetricValue"] {
        justify-content: center !important;
        font-family: "Trebuchet MS", Verdana, sans-serif !important;
        font-size: 3.35rem !important;
        line-height: 1.15 !important;
        font-weight: 700 !important;
        color: #b00020 !important;
    }

    [data-testid="stMetricValue"] div,
    [data-testid="stMetricValue"] span {
        font-family: "Trebuchet MS", Verdana, sans-serif !important;
    }

    @media print {
        h2 {
            break-before: page;
            page-break-before: always;
        }

        h2:first-of-type {
            break-before: auto;
            page-break-before: auto;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# EJECUCIÓN DEL NÚCLEO CON AJUSTES DE PRESENTACIÓN
# =========================================================

core_path = Path(__file__).with_name("_app_core.py")
core_source = core_path.read_text(encoding="utf-8")

core_source = core_source.replace(
    '''st.set_page_config(
    page_title="Evaluación Inicial Lengua 2.º ESO*",
    page_icon="📚",
    layout="centered",
)
''',
    "",
)

core_source = core_source.replace(
    "1.0 / 7",
    "1.0 / 6",
)

core_source = core_source.replace(
    '                    "exhortativa",\n',
    "",
)

# ---------------------------------------------------------
# 1. COMPRENSIÓN: preguntas 1-3 en rojo
# ---------------------------------------------------------
core_source = core_source.replace(
    '''        respuestas[pid] = st.text_input(
            pregunta["enunciado"],
            help=(
                "Escribe una respuesta breve."
                if pid != "c4"
                else
                "Escribe tres acciones separadas por comas."
            ),
            key=pid,
        )''',
    '''        if pid in ["c1", "c2", "c3"]:
            st.markdown(
                f'<div class="pregunta-roja">{pregunta["enunciado"]}</div>',
                unsafe_allow_html=True,
            )
            respuestas[pid] = st.text_input(
                "",
                help="Escribe una respuesta breve.",
                key=pid,
                label_visibility="collapsed",
            )
        else:
            respuestas[pid] = st.text_input(
                pregunta["enunciado"],
                help="Escribe tres acciones separadas por comas.",
                key=pid,
            )'''
)

# ---------------------------------------------------------
# 2. MORFOLOGÍA: indicar claramente qué debe escribir
# ---------------------------------------------------------
core_source = core_source.replace(
    '''            if campo == "Estructura":

                valor = st.selectbox(
                    f"{palabra['palabra']} → {campo}",''',
    '''            if campo == "Estructura":

                valor = st.selectbox(
                    "Escribe la estructura de " + palabra["palabra"],'''
)

core_source = core_source.replace(
    '''            elif campo == "V/I":

                valor = st.selectbox(
                    f"{palabra['palabra']} → {campo}",''',
    '''            elif campo == "V/I":

                valor = st.selectbox(
                    "Escribe si " + palabra["palabra"] + " es variable o invariable",'''
)

core_source = core_source.replace(
    '''            elif campo == "Categoría gramatical":

                valor = st.text_input(
                    f"{palabra['palabra']} → {campo}",''',
    '''            elif campo == "Categoría gramatical":

                valor = st.text_input(
                    "Escribe la categoría gramatical de " + palabra["palabra"],'''
)

core_source = core_source.replace(
    '''            else:

                valor = st.text_input(
                    f"{palabra['palabra']} → {campo}",''',
    '''            else:

                etiqueta_campo = {
                    "Lexema": "Escribe el lexema de ",
                    "Morfemas": "Escribe los morfemas de ",
                }.get(campo, f"{campo} de ")

                valor = st.text_input(
                    etiqueta_campo + palabra["palabra"],'''
)

# ---------------------------------------------------------
# 3. SEMÁNTICA: una sola aparición de las palabras dadas;
#    "relación semántica" queda resaltado en rojo.
# ---------------------------------------------------------
semantic_block_old = '''    for pregunta in EXAM[
        "semantica"
    ]:

        st.write(
            f"**{pregunta['elemento']}**"
        )

        respuestas[
            pregunta["id"]
        ] = st.selectbox(

            pregunta["enunciado"],

            opciones_semantica,

            key=pregunta["id"],
        )'''

semantic_block_new = '''    for pregunta in EXAM[
        "semantica"
    ]:

        elemento = pregunta["elemento"]
        enunciado = pregunta["enunciado"]

        # Solo mostramos una vez las palabras dadas.
        if "relación semántica" in enunciado:
            texto_pregunta = enunciado.split("→", 1)[-1].strip()
            texto_pregunta = texto_pregunta.replace(
                "¿qué relación semántica tienen?",
                '<span class="etiqueta-roja">¿qué relación semántica tienen?</span>'
            )
        else:
            texto_pregunta = enunciado.split("→", 1)[-1].strip()

        st.markdown(
            f'<div class="pregunta-roja">{elemento} → {texto_pregunta}</div>',
            unsafe_allow_html=True,
        )

        respuestas[
            pregunta["id"]
        ] = st.selectbox(
            "",
            opciones_semantica,
            key=pregunta["id"],
            label_visibility="collapsed",
        )'''

core_source = core_source.replace(
    semantic_block_old,
    semantic_block_new,
)

# También corregimos las dos formulaciones que había modificado el wrapper anterior.
core_source = core_source.replace(
    '"**Frío / calor** → ¿qué tipo de relación tienen?"',
    '"**Frío / calor** → ¿qué relación semántica tienen?"',
)
core_source = core_source.replace(
    '"**Perro, gato, caballo** → ¿qué tipo de relación forman?"',
    '"**Perro, gato, caballo** → ¿qué relación semántica tienen?"',
)

# ---------------------------------------------------------
# 4. TIPOS DE TEXTO: cada texto seguido inmediatamente de su pregunta.
# ---------------------------------------------------------
textos_block_old = '''    for letra, texto in EXAM[
        "textos"
    ]["textos"].items():

        st.markdown(
            f"**Texto {letra}:** {texto}"
        )

    for pregunta in EXAM[
        "textos"
    ]["preguntas"]:

        respuestas[
            pregunta["id"]
        ] = st.selectbox(

            pregunta["enunciado"],

            opciones_texto,

            key=pregunta["id"],
        )'''

textos_block_new = '''    textos = EXAM["textos"]["textos"]
    preguntas_textos = EXAM["textos"]["preguntas"]

    for indice, (letra, texto) in enumerate(textos.items()):

        st.markdown(
            f"**Texto {letra}:** {texto}"
        )

        if indice < len(preguntas_textos):
            pregunta = preguntas_textos[indice]
            respuestas[pregunta["id"]] = st.selectbox(
                pregunta["enunciado"],
                opciones_texto,
                key=pregunta["id"],
            )'''

core_source = core_source.replace(
    textos_block_old,
    textos_block_new,
)

# ---------------------------------------------------------
# 5. EXCEL: misma estructura de exportación individual que
#    el examen ordinario, adaptada a los campos NNEE.
# ---------------------------------------------------------
excel_function = r'''

def excel_individual(fila, respuestas, perfil):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()

    # HOJA 1: RESULTADO
    ws = wb.active
    ws.title = "Resultado"

    ws["A1"] = "📚 Evaluación inicial de Lengua — 2.º ESO"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "Alumno"
    ws["B3"] = fila["name"]
    ws["A4"] = "Grupo"
    ws["B4"] = fila["group"]
    ws["A5"] = "Fecha y hora"
    ws["B5"] = fila["date"]

    ws["A7"] = "NOTA FINAL (SOBRE 10)"
    ws["B7"] = fila["nota_final"]

    ws["A9"] = "RESULTADOS POR ÁREAS"
    ws["A9"].font = Font(bold=True)

    fila_excel = 10
    for clave, nombre_area in NOMBRES.items():
        ws.cell(fila_excel, 1).value = nombre_area
        ws.cell(fila_excel, 2).value = fila[clave]
        fila_excel += 1

    ws[f"A{fila_excel + 1}"] = "PERFIL COMPETENCIAL"
    ws[f"A{fila_excel + 1}"].font = Font(bold=True)

    fila_perfil = fila_excel + 2
    for item in perfil:
        ws.cell(fila_perfil, 1).value = item["texto"]
        ws.cell(fila_perfil, 1).alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )
        fila_perfil += 1

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 30

    # HOJA 2: RESPUESTAS
    ws2 = wb.create_sheet("Respuestas")
    ws2["A1"] = "Pregunta"
    ws2["B1"] = "Respuesta"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)

    fila_respuesta = 2
    for pregunta, respuesta in respuestas.items():
        ws2.cell(fila_respuesta, 1).value = pregunta
        ws2.cell(fila_respuesta, 2).value = str(respuesta)
        ws2.cell(fila_respuesta, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )
        fila_respuesta += 1

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 80

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida.getvalue()
'''

core_source = core_source.replace(
    '\n\ndef safe_read_results():',
    excel_function + '\n\ndef safe_read_results():',
)

# Añadimos el botón Excel justo después del CSV, igual que en el ordinario.
excel_download = '''

    st.download_button(
        "📊 Descargar Excel",
        data=excel_individual(
            fila,
            st.session_state.get("respuestas", {}),
            generar_perfil(scores),
        ),
        file_name=(
            "Resultado_2ESO_NEE_"
            + re.sub(
                r"[^A-Za-z0-9_-]+",
                "_",
                nombre_resultado,
            )
            + ".xlsx"
        ),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
'''

# Guardamos las respuestas para que estén disponibles al volver a ejecutar.
core_source = core_source.replace(
    '''    st.session_state[
        "fila"
    ] = fila

    st.rerun()''',
    '''    st.session_state[
        "fila"
    ] = fila

    st.session_state[
        "respuestas"
    ] = respuestas

    st.rerun()'''
)

# El bloque de resultados tiene el CSV del ordinario; insertamos Excel debajo.
marker = '''        use_container_width=True,
    )


    # =====================================================
    # ESTADÍSTICAS DEL GRUPO'''
core_source = core_source.replace(
    marker,
    '''        use_container_width=True,
    )
''' + excel_download + '''

    # =====================================================
    # ESTADÍSTICAS DEL GRUPO''',
)

core_namespace = {
    "__name__": "_app_core",
    "__file__": str(core_path),
}

exec(
    compile(core_source, str(core_path), "exec"),
    core_namespace,
)
